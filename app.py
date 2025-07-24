import streamlit as st
from openpyxl import Workbook, load_workbook
import os

# Excel file name
excel_file = "form_data.xlsx"

# Create Excel file with headers if not exists
if not os.path.exists(excel_file):
  wb = Workbook()
  ws = wb.active
  ws.append(["Name", "Age", "Gender", "Email", "Address"])
  wb.save(excel_file)

# Streamlit UI
st.title("📝 Data Entry Form")

with st.form("entry_form"):
  name = st.text_input("Name *")
  age = st.text_input("Age *")
  gender = st.selectbox("Gender", ["Select", "Male", "Female", "Other"])
  email = st.text_input("Email *")
  address = st.text_area("Address")

  submitted = st.form_submit_button("Submit")

  if submitted:
    if not name or not age or not email:
      st.error("Please fill all required fields (Name, Age, Email).")
    else:
      wb = load_workbook(excel_file)
      ws = wb.active
      ws.append([name, age, gender, email, address])
      wb.save(excel_file)
      st.success("✅ Your response has been recorded.")

# Optional: Download Excel
with open(excel_file, "rb") as f:
  st.download_button("📥 Download Excel File", f, file_name="form_data.xlsx")
